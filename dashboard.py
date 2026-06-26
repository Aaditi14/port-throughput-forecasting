import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, 
                              Border, Side)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ── Load Data ─────────────────────────────────────────────────────────────────
df = pd.read_csv("data/jnpt_throughput.csv")
df["date"] = pd.to_datetime(df["date"])

# ── Prepare Summary Tables ────────────────────────────────────────────────────

# 1. Annual Summary
annual = df.groupby("year").agg(
    total_teu    = ("teu_volume", "sum"),
    avg_monthly  = ("teu_volume", "mean"),
    min_monthly  = ("teu_volume", "min"),
    max_monthly  = ("teu_volume", "max")
).reset_index()
annual["total_teu_M"]   = (annual["total_teu"] / 1_000_000).round(3)
annual["avg_monthly"]   = annual["avg_monthly"].round(0).astype(int)
annual["yoy_growth_pct"] = annual["total_teu"].pct_change().mul(100).round(2)

# 2. Seasonality (avg TEU by calendar month)
seasonality = df.groupby(["month", "month_name"]).agg(
    avg_teu = ("teu_volume", "mean"),
    min_teu = ("teu_volume", "min"),
    max_teu = ("teu_volume", "max")
).reset_index()
seasonality["avg_teu"] = seasonality["avg_teu"].round(0).astype(int)
seasonality["min_teu"] = seasonality["min_teu"].astype(int)
seasonality["max_teu"] = seasonality["max_teu"].astype(int)

# 3. COVID Impact
covid = df[df["year"].isin([2019, 2020])].copy()
covid_pivot = covid.pivot_table(
    index=["month", "month_name"],
    columns="year",
    values="teu_volume",
    aggfunc="sum"
).reset_index()
covid_pivot.columns = ["month", "month_name", "teu_2019", "teu_2020"]
covid_pivot["change_pct"] = (
    (covid_pivot["teu_2020"] - covid_pivot["teu_2019"])
    / covid_pivot["teu_2019"] * 100
).round(2)

# ── Style Helpers ─────────────────────────────────────────────────────────────
BLUE_DARK   = "1F3864"
BLUE_MID    = "2E75B6"
BLUE_LIGHT  = "D6E4F0"
ORANGE      = "C55A11"
WHITE       = "FFFFFF"
GREY_LIGHT  = "F2F2F2"

def header_style(cell, bg=BLUE_DARK, fg=WHITE, size=11, bold=True):
    cell.font      = Font(bold=bold, color=fg, size=size, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)

def data_style(cell, bg=WHITE, bold=False, align="center"):
    cell.font      = Font(bold=bold, color="000000", size=10, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def write_df_to_sheet(ws, df, start_row=1, start_col=1,
                       header_bg=BLUE_DARK):
    """Write a dataframe to a worksheet with styling."""
    cols = list(df.columns)

    # Header row
    for ci, col in enumerate(cols, start=start_col):
        cell = ws.cell(row=start_row, column=ci, value=col.replace("_", " ").title())
        header_style(cell, bg=header_bg)
        cell.border = thin_border()

    # Data rows
    for ri, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for ci, val in enumerate(row, start=start_col):
            cell = ws.cell(row=ri, column=ci, value=val)
            bg   = GREY_LIGHT if (ri - start_row) % 2 == 0 else WHITE
            data_style(cell, bg=bg)
            cell.border = thin_border()

    # Auto-width columns
    for ci, col in enumerate(cols, start=start_col):
        max_len = max(len(str(col)), 
                      df.iloc[:, ci - start_col].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 25)

# ── Create Workbook ───────────────────────────────────────────────────────────
wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Cover / Summary
# ═══════════════════════════════════════════════════════════════════════════════
ws_cover = wb.active
ws_cover.title = "Dashboard"
ws_cover.sheet_view.showGridLines = False

# Title block
ws_cover.merge_cells("B2:H3")
title_cell = ws_cover["B2"]
title_cell.value = "JNPT Container Throughput Analysis Dashboard"
title_cell.font  = Font(bold=True, size=18, color=WHITE, name="Calibri")
title_cell.fill  = PatternFill("solid", fgColor=BLUE_DARK)
title_cell.alignment = Alignment(horizontal="center", vertical="center")

ws_cover.merge_cells("B4:H4")
sub_cell = ws_cover["B4"]
sub_cell.value = "Period: Jan 2015 – Dec 2024  |  Scale: ~1.5M – 2.1M TEUs/year  |  GTI – APM Terminals Mumbai"
sub_cell.font  = Font(bold=False, size=11, color=WHITE, name="Calibri")
sub_cell.fill  = PatternFill("solid", fgColor=BLUE_MID)
sub_cell.alignment = Alignment(horizontal="center", vertical="center")

ws_cover.row_dimensions[2].height = 35
ws_cover.row_dimensions[4].height = 22

# KPI boxes
kpis = [
    ("Total Records",    "120 Months",       "B"),
    ("Period Covered",   "2015 – 2024",      "D"),
    ("Peak Annual TEU",  "2.08M (2024)",     "F"),
    ("Avg YoY Growth",   "~3.7%",            "H"),
]
for label, value, col in kpis:
    ws_cover.merge_cells(f"{col}6:{col}7")
    ws_cover.merge_cells(f"{col}8:{col}9")
    lc = ws_cover[f"{col}6"]
    vc = ws_cover[f"{col}8"]
    lc.value = label
    vc.value = value
    lc.font  = Font(bold=True,  size=10, color=WHITE,   name="Calibri")
    vc.font  = Font(bold=True,  size=14, color=BLUE_DARK, name="Calibri")
    lc.fill  = PatternFill("solid", fgColor=BLUE_MID)
    vc.fill  = PatternFill("solid", fgColor=BLUE_LIGHT)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    vc.alignment = Alignment(horizontal="center", vertical="center")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Annual Summary + Line Chart
# ═══════════════════════════════════════════════════════════════════════════════
ws_annual = wb.create_sheet("Annual Summary")
ws_annual.sheet_view.showGridLines = False

ws_annual.merge_cells("A1:G1")
h = ws_annual["A1"]
h.value = "Annual Throughput Summary (TEUs)"
header_style(h, bg=BLUE_DARK, size=13)
ws_annual.row_dimensions[1].height = 28

write_df_to_sheet(ws_annual, annual, start_row=2)

# Line chart — Total TEU by Year
chart1 = LineChart()
chart1.title   = "Annual Total TEU Volume (2015–2024)"
chart1.y_axis.title = "Total TEU"
chart1.x_axis.title = "Year"
chart1.style   = 10
chart1.width   = 22
chart1.height  = 14

data_ref = Reference(ws_annual, min_col=2, max_col=2,
                     min_row=2, max_row=len(annual) + 2)
cats_ref = Reference(ws_annual, min_col=1,
                     min_row=3, max_row=len(annual) + 2)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.series[0].graphicalProperties.line.solidFill = BLUE_MID
chart1.series[0].graphicalProperties.line.width     = 25000
ws_annual.add_chart(chart1, "A14")

# Bar chart — YoY Growth %
chart2 = BarChart()
chart2.title   = "Year-over-Year Growth Rate (%)"
chart2.y_axis.title = "Growth %"
chart2.x_axis.title = "Year"
chart2.style   = 10
chart2.width   = 22
chart2.height  = 14

yoy_col = annual.columns.get_loc("yoy_growth_pct") + 1
data_ref2 = Reference(ws_annual, min_col=yoy_col, max_col=yoy_col,
                      min_row=2, max_row=len(annual) + 2)
chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cats_ref)
chart2.series[0].graphicalProperties.solidFill = ORANGE
ws_annual.add_chart(chart2, "K14")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Seasonality + Bar Chart
# ═══════════════════════════════════════════════════════════════════════════════
ws_season = wb.create_sheet("Seasonality")
ws_season.sheet_view.showGridLines = False

ws_season.merge_cells("A1:F1")
h2 = ws_season["A1"]
h2.value = "Seasonality Analysis — Average TEU by Calendar Month"
header_style(h2, bg=BLUE_DARK, size=13)
ws_season.row_dimensions[1].height = 28

write_df_to_sheet(ws_season, seasonality, start_row=2)

# Bar chart — Avg TEU by Month
chart3 = BarChart()
chart3.title        = "Average Monthly TEU Volume (Seasonality)"
chart3.y_axis.title = "Avg TEU"
chart3.x_axis.title = "Month"
chart3.style        = 10
chart3.width        = 22
chart3.height       = 14

data_ref3 = Reference(ws_season, min_col=3, max_col=3,
                      min_row=2, max_row=len(seasonality) + 2)
cats_ref3 = Reference(ws_season, min_col=2,
                      min_row=3, max_row=len(seasonality) + 2)
chart3.add_data(data_ref3, titles_from_data=True)
chart3.set_categories(cats_ref3)
chart3.series[0].graphicalProperties.solidFill = BLUE_MID
ws_season.add_chart(chart3, "A15")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Raw Data
# ═══════════════════════════════════════════════════════════════════════════════
ws_raw = wb.create_sheet("Raw Data")
ws_raw.sheet_view.showGridLines = False

ws_raw.merge_cells("A1:H1")
h3 = ws_raw["A1"]
h3.value = "Raw Monthly Throughput Data — JNPT (Jan 2015 – Dec 2024)"
header_style(h3, bg=BLUE_DARK, size=13)
ws_raw.row_dimensions[1].height = 28

write_df_to_sheet(ws_raw, df, start_row=2)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — COVID Impact
# ═══════════════════════════════════════════════════════════════════════════════
ws_covid = wb.create_sheet("COVID Impact")
ws_covid.sheet_view.showGridLines = False

ws_covid.merge_cells("A1:E1")
h4 = ws_covid["A1"]
h4.value = "COVID-19 Impact — 2019 vs 2020 Monthly Comparison"
header_style(h4, bg=ORANGE, size=13)
ws_covid.row_dimensions[1].height = 28

write_df_to_sheet(ws_covid, covid_pivot, start_row=2, header_bg=ORANGE)

# Bar chart — 2019 vs 2020
chart4 = BarChart()
chart4.title        = "2019 vs 2020 Monthly TEU — COVID Impact"
chart4.y_axis.title = "TEU Volume"
chart4.x_axis.title = "Month"
chart4.style        = 10
chart4.width        = 22
chart4.height       = 14
chart4.grouping     = "clustered"

data_ref4 = Reference(ws_covid, min_col=3, max_col=4,
                      min_row=2, max_row=len(covid_pivot) + 2)
cats_ref4 = Reference(ws_covid, min_col=2,
                      min_row=3, max_row=len(covid_pivot) + 2)
chart4.add_data(data_ref4, titles_from_data=True)
chart4.set_categories(cats_ref4)
chart4.series[0].graphicalProperties.solidFill = BLUE_MID
chart4.series[1].graphicalProperties.solidFill = ORANGE
ws_covid.add_chart(chart4, "A15")

# ── Save ──────────────────────────────────────────────────────────────────────
output_path = "outputs/throughput_dashboard.xlsx"
wb.save(output_path)

print("=" * 55)
print(f"✅ Excel Dashboard saved → {output_path}")
print(f"   Sheets created : 5")
print(f"   Charts created : 4")
print(f"   Rows of data   : {len(df)}")
print("=" * 55)
print("\nSheets:")
print("  1. Dashboard     — KPI summary cover page")
print("  2. Annual Summary — TEU totals + YoY growth charts")
print("  3. Seasonality   — Avg TEU by month + bar chart")
print("  4. Raw Data      — Full 120-row dataset")
print("  5. COVID Impact  — 2019 vs 2020 comparison chart")